# app.py

import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import hashlib
import hmac
from supabase import create_client

# =========================================================
# PAGE CONFIG
# =========================================================

st.set_page_config(
    page_title="File Registration",
    layout="wide"
)

st.title("📂 File Registration Generator")

# =========================================================
# LOGIN SECURITY USING secrets.toml
# =========================================================

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if "current_user" not in st.session_state:
    st.session_state.current_user = None

if "current_role" not in st.session_state:
    st.session_state.current_role = None


def check_password(username, password):
    try:
        allowed_users = dict(st.secrets["users"])
    except Exception:
        st.error("❌ Missing [users] section in .streamlit/secrets.toml")
        return False

    if username not in allowed_users:
        return False

    correct_password = str(allowed_users[username])

    return hmac.compare_digest(password, correct_password)


def get_user_role(username):
    try:
        roles = dict(st.secrets["roles"])
        return roles.get(username, "staff")
    except Exception:
        return "staff"


def login_area():
    st.subheader("🔐 Login Required")

    username = st.text_input("Username")
    password = st.text_input("Password", type="password")

    login_button = st.button("Login")

    if login_button:

        if check_password(username, password):

            st.session_state.logged_in = True
            st.session_state.current_user = username
            st.session_state.current_role = get_user_role(username)

            st.success("✅ Login successful")
            st.rerun()

        else:

            st.error("❌ Invalid username or password")


def logout_area():
    col1, col2 = st.columns([4, 1])

    with col1:
        st.info(
            f"Logged in as: {st.session_state.current_user} "
            f"({st.session_state.current_role})"
        )

    with col2:
        if st.button("Logout"):
            st.session_state.logged_in = False
            st.session_state.current_user = None
            st.session_state.current_role = None

            if "records" in st.session_state:
                del st.session_state.records

            st.rerun()


def is_admin():
    return st.session_state.current_role == "admin"


if not st.session_state.logged_in:
    login_area()
    st.stop()
else:
    logout_area()

# =========================================================
# HELPER FUNCTIONS
# =========================================================

def clean_text(value):
    if pd.isna(value):
        return ""

    return str(value).replace("\r", "").strip()


def safe_get_row_value(row, position):
    try:
        return row[position]
    except Exception:
        return ""


def split_solicitor_values(value_text):
    if pd.isna(value_text):
        value_text = ""
    else:
        value_text = str(value_text).replace("\r", "")

    value_lines = [
        line.strip()
        for line in value_text.split("\n")
    ]

    while len(value_lines) < 5:
        value_lines.append("")

    return {
        "V Solicitor": value_lines[0],
        "V Financier": value_lines[1],
        "P Solicitor": value_lines[2],
        "P Financier": value_lines[3],
        "B Solicitor": value_lines[4]
    }


def safe_date_for_input(value):
    parsed_date = pd.to_datetime(
        value,
        errors="coerce"
    )

    if pd.isna(parsed_date):
        return pd.Timestamp.today().date()

    return parsed_date.date()


# =========================================================
# SUPABASE DATABASE FUNCTIONS
# =========================================================

@st.cache_resource
def get_supabase_client():
    try:
        return create_client(
            st.secrets["supabase"]["url"],
            st.secrets["supabase"]["key"]
        )
    except Exception:
        st.error("❌ Missing [supabase] section in .streamlit/secrets.toml")
        st.stop()


supabase = get_supabase_client()


def db_row_to_record(row):
    return {
        "ID": row.get("id"),
        "Date": pd.to_datetime(row.get("date"), errors="coerce"),
        "File Reference": row.get("file_reference", ""),
        "Vendors": row.get("vendors", ""),
        "Purchasers": row.get("purchasers", ""),
        "Property": row.get("property", ""),
        "Purchase Price": row.get("purchase_price", ""),
        "V Solicitor": row.get("v_solicitor", ""),
        "V Financier": row.get("v_financier", ""),
        "P Solicitor": row.get("p_solicitor", ""),
        "P Financier": row.get("p_financier", ""),
        "B Solicitor": row.get("b_solicitor", "")
    }


def record_to_db_row(record):
    date_value = record.get("Date")

    if pd.notna(date_value):
        date_value = pd.to_datetime(date_value).strftime("%Y-%m-%d")
    else:
        date_value = None

    return {
        "date": date_value,
        "file_reference": record.get("File Reference", ""),
        "vendors": record.get("Vendors", ""),
        "purchasers": record.get("Purchasers", ""),
        "property": record.get("Property", ""),
        "purchase_price": record.get("Purchase Price", ""),
        "v_solicitor": record.get("V Solicitor", ""),
        "v_financier": record.get("V Financier", ""),
        "p_solicitor": record.get("P Solicitor", ""),
        "p_financier": record.get("P Financier", ""),
        "b_solicitor": record.get("B Solicitor", ""),
        "created_by": st.session_state.current_user
    }


def load_records_from_db():
    try:
        response = (
            supabase
            .table("file_records")
            .select("*")
            .order("id", desc=False)
            .execute()
        )

        return [
            db_row_to_record(row)
            for row in response.data
        ]

    except Exception as e:
        st.error(f"❌ Database Load Error: {e}")
        return []


def add_record_to_db(record):
    try:
        response = (
            supabase
            .table("file_records")
            .insert(record_to_db_row(record))
            .execute()
        )

        return response

    except Exception as e:
        st.error(f"❌ Database Add Error: {e}")
        return None


def update_record_in_db(record_id, record):
    try:
        response = (
            supabase
            .table("file_records")
            .update(record_to_db_row(record))
            .eq("id", record_id)
            .execute()
        )

        return response

    except Exception as e:
        st.error(f"❌ Database Update Error: {e}")
        return None


def delete_record_from_db(record_id):
    try:
        response = (
            supabase
            .table("file_records")
            .delete()
            .eq("id", record_id)
            .execute()
        )

        return response

    except Exception as e:
        st.error(f"❌ Database Delete Error: {e}")
        return None


def delete_all_records_from_db():
    try:
        (
            supabase
            .table("file_records")
            .delete()
            .neq("id", -1)
            .execute()
        )

    except Exception as e:
        st.error(f"❌ Database Clear Error: {e}")


def import_records_to_db(records):
    try:
        delete_all_records_from_db()

        rows = [
            record_to_db_row(record)
            for record in records
        ]

        if rows:
            (
                supabase
                .table("file_records")
                .insert(rows)
                .execute()
            )

    except Exception as e:
        st.error(f"❌ Database Import Error: {e}")


# =========================================================
# SESSION STATE
# =========================================================

if "records" not in st.session_state:
    st.session_state.records = load_records_from_db()

if "uploaded_file_id" not in st.session_state:
    st.session_state.uploaded_file_id = None

if "file_imported" not in st.session_state:
    st.session_state.file_imported = False

# This is used to clear the add form after successful submission.
if "add_form_version" not in st.session_state:
    st.session_state.add_form_version = 0


# =========================================================
# REFRESH BUTTON
# =========================================================

if st.button("🔄 Refresh Records"):
    st.session_state.records = load_records_from_db()
    st.success("✅ Records refreshed from database")
    st.rerun()


# =========================================================
# IMPORT EXISTING EXCEL FILE
# EXACT FORMAT:
# A No
# B Date
# C File Reference
# D V/P
# E Client(s) Particulars
# F Property
# G Purchase Price
# H R
# I Solicitor Labels
# J Solicitor Values
# =========================================================

if is_admin():

    st.subheader("📤 Import Existing Excel File")

    st.warning(
        "⚠️ Uploading an Excel file will replace all records currently saved in Supabase."
    )

    uploaded_file = st.file_uploader(
        "Upload existing Excel file",
        type=["xlsx"]
    )

    if uploaded_file is not None:

        uploaded_file_bytes = uploaded_file.getvalue()

        current_file_id = hashlib.md5(
            uploaded_file_bytes
        ).hexdigest()

        if (
            st.session_state.uploaded_file_id != current_file_id
            or st.session_state.file_imported is False
        ):

            try:

                raw_df = pd.read_excel(
                    BytesIO(uploaded_file_bytes),
                    header=None
                )

                imported_records = []

                row_index = 2

                while row_index < len(raw_df):

                    row_top = raw_df.iloc[row_index]

                    file_ref = clean_text(
                        safe_get_row_value(row_top, 2)
                    )

                    if not file_ref or file_ref.lower() == "nan":

                        row_index += 1
                        continue

                    file_date = pd.to_datetime(
                        safe_get_row_value(row_top, 1),
                        errors="coerce"
                    )

                    vendors = clean_text(
                        safe_get_row_value(row_top, 4)
                    )

                    purchasers = ""

                    if row_index + 1 < len(raw_df):

                        row_bottom = raw_df.iloc[row_index + 1]

                        bottom_file_ref = clean_text(
                            safe_get_row_value(row_bottom, 2)
                        )

                        bottom_vp_label = clean_text(
                            safe_get_row_value(row_bottom, 3)
                        )

                        bottom_client = clean_text(
                            safe_get_row_value(row_bottom, 4)
                        )

                        if (
                            not bottom_file_ref
                            and bottom_vp_label.upper().startswith("P")
                        ):

                            purchasers = bottom_client
                            row_index += 1

                    property_name = clean_text(
                        safe_get_row_value(row_top, 5)
                    )

                    purchase_price = clean_text(
                        safe_get_row_value(row_top, 6)
                    )

                    solicitor_value_cell = safe_get_row_value(row_top, 9)

                    solicitor_values = split_solicitor_values(
                        solicitor_value_cell
                    )

                    imported_records.append({

                        "Date": file_date,

                        "File Reference": file_ref,

                        "Vendors": vendors,

                        "Purchasers": purchasers,

                        "Property": property_name,

                        "Purchase Price": purchase_price,

                        "V Solicitor": solicitor_values["V Solicitor"],

                        "V Financier": solicitor_values["V Financier"],

                        "P Solicitor": solicitor_values["P Solicitor"],

                        "P Financier": solicitor_values["P Financier"],

                        "B Solicitor": solicitor_values["B Solicitor"]
                    })

                    row_index += 1

                import_records_to_db(imported_records)

                st.session_state.records = load_records_from_db()

                st.session_state.uploaded_file_id = current_file_id

                st.session_state.file_imported = True

                st.success(
                    f"✅ Imported {len(imported_records)} records successfully into Supabase!"
                )

                st.rerun()

            except Exception as e:

                st.error(f"Import Error: {e}")

        else:

            st.info(
                "✅ Excel file already imported. You can continue adding records."
            )

else:

    st.info("📤 Excel import is only available for admin users.")


# =========================================================
# ADD RECORD FORM
# =========================================================

form_version = st.session_state.add_form_version

with st.form("registration_form", enter_to_submit=False):

    st.subheader("➕ Add New Record")

    col1, col2 = st.columns(2)

    with col1:

        date = st.date_input(
            "Date",
            key=f"add_date_{form_version}"
        )

        file_ref = st.text_input(
            "File Reference",
            key=f"add_file_ref_{form_version}"
        )

    with col2:

        property_name = st.text_area(
            "Property",
            key=f"add_property_{form_version}"
        )

        purchase_price = st.text_input(
            "Purchase Price",
            key=f"add_purchase_price_{form_version}"
        )

    st.subheader("Client(s) Particulars")

    col3, col4 = st.columns(2)

    with col3:

        st.markdown("### Vendors")

        vendor1 = st.text_input(
            "Vendor 1",
            key=f"add_vendor1_{form_version}"
        )

        vendor2 = st.text_input(
            "Vendor 2",
            key=f"add_vendor2_{form_version}"
        )

        vendor3 = st.text_input(
            "Vendor 3",
            key=f"add_vendor3_{form_version}"
        )

    with col4:

        st.markdown("### Purchasers")

        purchaser1 = st.text_input(
            "Purchaser 1",
            key=f"add_purchaser1_{form_version}"
        )

        purchaser2 = st.text_input(
            "Purchaser 2",
            key=f"add_purchaser2_{form_version}"
        )

        purchaser3 = st.text_input(
            "Purchaser 3",
            key=f"add_purchaser3_{form_version}"
        )

    st.subheader("Solicitor / Financier")

    col5, col6 = st.columns(2)

    with col5:

        v_solicitor = st.text_input(
            "V Solicitor",
            key=f"add_v_solicitor_{form_version}"
        )

        v_financier = st.text_input(
            "V Financier",
            key=f"add_v_financier_{form_version}"
        )

    with col6:

        p_solicitor = st.text_input(
            "P Solicitor",
            key=f"add_p_solicitor_{form_version}"
        )

        p_financier = st.text_input(
            "P Financier",
            key=f"add_p_financier_{form_version}"
        )

        b_solicitor = st.text_input(
            "B Solicitor",
            key=f"add_b_solicitor_{form_version}"
        )

    submitted = st.form_submit_button("Add Record")

    if submitted:

        vendors = "\n".join([
            v for v in [
                vendor1,
                vendor2,
                vendor3
            ] if v.strip()
        ])

        purchasers = "\n".join([
            p for p in [
                purchaser1,
                purchaser2,
                purchaser3
            ] if p.strip()
        ])

        new_record = {

            "Date": pd.to_datetime(date),

            "File Reference": file_ref,

            "Vendors": vendors,

            "Purchasers": purchasers,

            "Property": property_name,

            "Purchase Price": purchase_price,

            "V Solicitor": v_solicitor,

            "V Financier": v_financier,

            "P Solicitor": p_solicitor,

            "P Financier": p_financier,

            "B Solicitor": b_solicitor
        }

        add_result = add_record_to_db(new_record)

        if add_result is not None:

            st.session_state.records = load_records_from_db()

            # This makes the form create a fresh blank set of fields.
            st.session_state.add_form_version += 1

            st.success("✅ Record Added Successfully")

            st.rerun()


# =========================================================
# DISPLAY TABLE / EDIT / DELETE / DOWNLOAD
# ONLY ADMIN CAN SEE THIS SECTION
# =========================================================

if st.session_state.records:

    if not is_admin():

        st.warning(
            "⚠️ You do not have permission to view the preview, edit records, delete records, or download the Excel file."
        )

        st.info(
            "You can still add new records, but only admin users can view the full file list."
        )

    else:

        df = pd.DataFrame(
            st.session_state.records
        )

        df["Date Sort"] = pd.to_datetime(
            df["Date"],
            errors="coerce"
        )

        # Keep table in normal order.
        # Older records are above, newer records are lower.
        df = df.sort_values(
            by=["Date Sort", "ID"],
            ascending=[True, True],
            na_position="last"
        )

        df_display = df.copy()

        df_display.insert(
            0,
            "No",
            range(1, len(df_display) + 1)
        )

        df_for_export = df.copy()

        if "ID" in df_for_export.columns:
            df_for_export = df_for_export.drop(columns=["ID"])

        if "Date Sort" in df_for_export.columns:
            df_for_export = df_for_export.drop(columns=["Date Sort"])

        # =====================================================
        # FOCUSED PREVIEW AROUND LATEST ENTRY
        # =====================================================

        st.subheader("📋 Preview - Focused Around Latest Entry")

        total_rows = len(df_display)

        st.info(f"📊 Total records in database: {total_rows}")

        st.caption(
            "Showing a small part of the table focused around the latest added record."
        )

        latest_record_id = df_display["ID"].max()

        latest_index_list = df_display.index[
            df_display["ID"] == latest_record_id
        ].tolist()

        if latest_index_list:

            latest_index = latest_index_list[0]

            latest_position = df_display.index.get_loc(latest_index)

            start_position = max(latest_position - 5, 0)

            end_position = min(latest_position + 6, len(df_display))

            focused_df = df_display.iloc[
                start_position:end_position
            ].copy()

        else:

            latest_index = None

            focused_df = df_display.head(10).copy()

        search_name = st.text_input(
            "🔍 Find by Vendor / Purchaser Name / File Reference"
        )


        def remove_internal_columns(dataframe):
            cleaned_df = dataframe.copy()

            if "ID" in cleaned_df.columns:
                cleaned_df = cleaned_df.drop(columns=["ID"])

            if "Date Sort" in cleaned_df.columns:
                cleaned_df = cleaned_df.drop(columns=["Date Sort"])

            return cleaned_df


        def highlight_latest_row(row):
            if row.name == latest_index:
                return [
                    "background-color: #fff3b0; color: black; font-weight: bold"
                ] * len(row)

            return [
                "color: white"
            ] * len(row)


        if search_name:

            filtered_df = df_display[
                df_display["Vendors"].str.contains(
                    search_name,
                    case=False,
                    na=False
                )
                |
                df_display["Purchasers"].str.contains(
                    search_name,
                    case=False,
                    na=False
                )
                |
                df_display["File Reference"].str.contains(
                    search_name,
                    case=False,
                    na=False
                )
            ]

            filtered_preview = remove_internal_columns(
                filtered_df.head(10)
            )

            if not filtered_preview.empty:

                styled_search_df = filtered_preview.style.apply(
                    highlight_latest_row,
                    axis=1
                )

                st.dataframe(
                    styled_search_df,
                    use_container_width=True
                )

            else:

                st.warning("No matching records found.")

        else:

            focused_preview = remove_internal_columns(focused_df)

            styled_focused_df = focused_preview.style.apply(
                highlight_latest_row,
                axis=1
            )

            st.dataframe(
                styled_focused_df,
                use_container_width=True
            )

        # =====================================================
        # SELECT RECORD
        # =====================================================

        st.subheader("✏️ Edit Existing Record")

        record_numbers = df_display["No"].tolist()

        selected_no = st.selectbox(
            "Select Record Number",
            record_numbers
        )

        selected_row = df_display[
            df_display["No"] == selected_no
        ].iloc[0]

        selected_record_id = selected_row["ID"]

        selected_record_matches = [
            record
            for record in st.session_state.records
            if record["ID"] == selected_record_id
        ]

        if not selected_record_matches:
            st.error("Selected record could not be found. Please refresh records.")
            st.stop()

        selected_record = selected_record_matches[0]

        # =====================================================
        # EDIT RECORD FORM
        # =====================================================

        with st.form("edit_form", enter_to_submit=False):

            edit_date = st.date_input(
                "Edit Date",
                value=safe_date_for_input(
                    selected_record["Date"]
                )
            )

            edit_file_ref = st.text_input(
                "Edit File Reference",
                value=selected_record["File Reference"]
            )

            edit_property = st.text_area(
                "Edit Property",
                value=selected_record["Property"]
            )

            edit_purchase_price = st.text_input(
                "Edit Purchase Price",
                value=selected_record.get("Purchase Price", "")
            )

            vendor_lines = (
                selected_record["Vendors"].split("\n")
                if selected_record["Vendors"]
                else []
            )

            purchaser_lines = (
                selected_record["Purchasers"].split("\n")
                if selected_record["Purchasers"]
                else []
            )

            col1, col2 = st.columns(2)

            with col1:

                st.markdown("### Vendors")

                edit_vendor1 = st.text_input(
                    "Edit Vendor 1",
                    value=vendor_lines[0]
                    if len(vendor_lines) > 0 else ""
                )

                edit_vendor2 = st.text_input(
                    "Edit Vendor 2",
                    value=vendor_lines[1]
                    if len(vendor_lines) > 1 else ""
                )

                edit_vendor3 = st.text_input(
                    "Edit Vendor 3",
                    value=vendor_lines[2]
                    if len(vendor_lines) > 2 else ""
                )

            with col2:

                st.markdown("### Purchasers")

                edit_purchaser1 = st.text_input(
                    "Edit Purchaser 1",
                    value=purchaser_lines[0]
                    if len(purchaser_lines) > 0 else ""
                )

                edit_purchaser2 = st.text_input(
                    "Edit Purchaser 2",
                    value=purchaser_lines[1]
                    if len(purchaser_lines) > 1 else ""
                )

                edit_purchaser3 = st.text_input(
                    "Edit Purchaser 3",
                    value=purchaser_lines[2]
                    if len(purchaser_lines) > 2 else ""
                )

            col3, col4 = st.columns(2)

            with col3:

                edit_v_solicitor = st.text_input(
                    "Edit V Solicitor",
                    value=selected_record["V Solicitor"]
                )

                edit_v_financier = st.text_input(
                    "Edit V Financier",
                    value=selected_record["V Financier"]
                )

            with col4:

                edit_p_solicitor = st.text_input(
                    "Edit P Solicitor",
                    value=selected_record["P Solicitor"]
                )

                edit_p_financier = st.text_input(
                    "Edit P Financier",
                    value=selected_record["P Financier"]
                )

                edit_b_solicitor = st.text_input(
                    "Edit B Solicitor",
                    value=selected_record["B Solicitor"]
                )

            update_button = st.form_submit_button(
                "Update Record"
            )

            if update_button:

                updated_vendors = "\n".join([
                    v for v in [
                        edit_vendor1,
                        edit_vendor2,
                        edit_vendor3
                    ] if v.strip()
                ])

                updated_purchasers = "\n".join([
                    p for p in [
                        edit_purchaser1,
                        edit_purchaser2,
                        edit_purchaser3
                    ] if p.strip()
                ])

                updated_record = {

                    "Date": pd.to_datetime(edit_date),

                    "File Reference": edit_file_ref,

                    "Vendors": updated_vendors,

                    "Purchasers": updated_purchasers,

                    "Property": edit_property,

                    "Purchase Price": edit_purchase_price,

                    "V Solicitor": edit_v_solicitor,

                    "V Financier": edit_v_financier,

                    "P Solicitor": edit_p_solicitor,

                    "P Financier": edit_p_financier,

                    "B Solicitor": edit_b_solicitor
                }

                update_result = update_record_in_db(
                    selected_record_id,
                    updated_record
                )

                if update_result is not None:

                    st.session_state.records = load_records_from_db()

                    st.success(
                        "✅ Record Updated Successfully"
                    )

                    st.rerun()

        # =====================================================
        # DELETE RECORD
        # =====================================================

        st.subheader("🗑️ Delete Selected Record")

        st.warning(
            f"Delete record No. {selected_no}: {selected_record['File Reference']}"
        )

        confirm_delete = st.checkbox(
            "I confirm I want to delete this whole row"
        )

        if st.button("🗑️ Delete This Record", type="primary"):

            if confirm_delete:

                delete_result = delete_record_from_db(selected_record_id)

                if delete_result is not None:

                    st.session_state.records = load_records_from_db()

                    st.success("✅ Record Deleted Successfully")

                    st.rerun()

            else:

                st.error("Please tick the confirmation checkbox before deleting.")

        # =====================================================
        # EXCEL GENERATOR
        # =====================================================

        def generate_excel(dataframe):

            wb = Workbook()

            ws = wb.active

            ws.title = "File Registration"

            ws.merge_cells("A1:J1")

            ws["A1"] = "File Registration"

            ws["A1"].font = Font(
                bold=True,
                size=12
            )

            headers = [
                "No",
                "Date",
                "File Reference",
                "",
                "Client(s) Particulars",
                "Property",
                "Purchase Price",
                "R",
                "",
                ""
            ]

            for col_num, header in enumerate(headers, 1):

                cell = ws.cell(
                    row=2,
                    column=col_num
                )

                cell.value = header

                cell.font = Font(bold=True)

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            widths = {
                1: 6,
                2: 12,
                3: 28,
                4: 5,
                5: 40,
                6: 40,
                7: 18,
                8: 8,
                9: 20,
                10: 25
            }

            for col, width in widths.items():

                ws.column_dimensions[
                    get_column_letter(col)
                ].width = width

            start_row = 3

            dataframe_export = dataframe.copy()

            dataframe_export = dataframe_export.sort_values(
                by="Date",
                ascending=True,
                na_position="last"
            )

            dataframe_export = dataframe_export.reset_index(
                drop=True
            )

            dataframe_export.insert(
                0,
                "No",
                range(1, len(dataframe_export) + 1)
            )

            for index, row in dataframe_export.iterrows():

                top_row = start_row
                bottom_row = start_row + 1

                merge_columns = [
                    1,
                    2,
                    3,
                    6,
                    7,
                    8,
                    9,
                    10
                ]

                for col in merge_columns:

                    ws.merge_cells(
                        start_row=top_row,
                        start_column=col,
                        end_row=bottom_row,
                        end_column=col
                    )

                ws.cell(top_row, 1).value = row["No"]

                if pd.notna(row["Date"]):

                    ws.cell(top_row, 2).value = (
                        pd.to_datetime(row["Date"]).strftime("%d/%m/%y")
                    )

                else:

                    ws.cell(top_row, 2).value = ""

                ws.cell(top_row, 3).value = row["File Reference"]

                ws.cell(top_row, 4).value = "V"

                ws.cell(bottom_row, 4).value = "P"

                ws.cell(top_row, 5).value = row["Vendors"]

                ws.cell(bottom_row, 5).value = row["Purchasers"]

                ws.cell(top_row, 6).value = row["Property"]

                ws.cell(top_row, 7).value = row.get("Purchase Price", "")

                label_text = (
                    "V Solicitor\n"
                    "V Financier\n"
                    "P Solicitor\n"
                    "P Financier\n"
                    "B Solicitor"
                )

                ws.cell(top_row, 9).value = label_text

                value_text = (
                    f'{row["V Solicitor"]}\n'
                    f'{row["V Financier"]}\n'
                    f'{row["P Solicitor"]}\n'
                    f'{row["P Financier"]}\n'
                    f'{row["B Solicitor"]}'
                )

                ws.cell(top_row, 10).value = value_text

                for r in [top_row, bottom_row]:

                    for c in range(1, 11):

                        cell = ws.cell(r, c)

                        cell.alignment = Alignment(
                            horizontal="left",
                            vertical="top",
                            wrap_text=True
                        )

                ws.row_dimensions[top_row].height = 45
                ws.row_dimensions[bottom_row].height = 45

                start_row += 2

            thin = Side(style="thin")

            for row in ws.iter_rows(
                min_row=1,
                max_row=ws.max_row,
                min_col=1,
                max_col=10
            ):

                for cell in row:

                    cell.border = Border(
                        left=thin,
                        right=thin,
                        top=thin,
                        bottom=thin
                    )

            for row in ws.iter_rows():

                for cell in row:

                    cell.font = Font(size=10)

            for cell in ws[2]:

                cell.font = Font(
                    bold=True,
                    size=10
                )

            output = BytesIO()

            wb.save(output)

            output.seek(0)

            return output

        excel_file = generate_excel(df_for_export)

        st.download_button(
            label="📥 Download Excel File",
            data=excel_file,
            file_name="file_registration.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

else:

    st.info("No records yet. Add a new record or import an Excel file.")